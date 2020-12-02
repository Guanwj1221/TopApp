#!/usr/bin/python3
# coding=utf-8
import re
# pip  install openpyxl
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.workbook.child import INVALID_TITLE_REGEX


# 判断一个List的维度
def list_dim(list_data):
    try:
        i = list_data[0][0][0]
        dim = 3
    except:
        try:
            i = list_data[0][0]
            dim = 2
        except:
            dim = 1
    return dim


# 讲内容写进Excel表格
def write_content(path, ranking_sheet_name,
                  ranking_content, reviews_sheet_names, reviews_content):
    try:
        workbook = openpyxl.load_workbook(path)
    except:
        workbook = openpyxl.Workbook()
    sheet_name = format_sheet_name(ranking_sheet_name)
    try:
        sheet = workbook[sheet_name]
    except:
        sheet = workbook.create_sheet(sheet_name)
    for i in range(len(ranking_content)):
        for j in range(0, len(ranking_content[i])):
            tem_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(ranking_content[i][j]))
            sheet.cell(row=i + 1, column=j + 1, value=tem_value)
    for index in range(len(reviews_content)):
        values = reviews_content[index]
        if values is None:
            continue
        sheet_name = format_sheet_name(reviews_sheet_names[index])
        try:
            sheet = workbook[sheet_name]
        except:
            sheet = workbook.create_sheet(sheet_name)
        for i in range(0, len(values)):
            for j in range(0, len(values[i])):
                tem_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(values[i][j]))
                sheet.cell(row=i + 1, column=j + 1, value=tem_value)
    try:
        if workbook.sheetnames[0] == 'Sheet':
            sheet = workbook[workbook.sheetnames[0]]
            workbook.remove(sheet)
    except:
        pass
    workbook.save(path)
    workbook.close()


def write_reviews(path, app_name, content):
    try:
        workbook = openpyxl.load_workbook(path)
    except:
        workbook = openpyxl.Workbook()
    sheet_name = format_sheet_name(app_name)
    try:
        sheet = workbook[sheet_name]
    except:
        sheet = workbook.create_sheet(sheet_name)
    for i in range(len(content)):
        for j in range(0, len(content[i])):
            tem_value = ILLEGAL_CHARACTERS_RE.sub(r'', str(content[i][j]))
            sheet.cell(row=i + 1, column=j + 1, value=tem_value)
    try:
        if workbook.sheetnames[0] == 'Sheet':
            sheet = workbook[workbook.sheetnames[0]]
            workbook.remove(sheet)
    except:
        pass
    workbook.save(path)
    workbook.close()


# 格式化sheet name
def format_sheet_name(sheet_name):
    name = str(sheet_name).replace(' ', '')
    name = re.sub(INVALID_TITLE_REGEX, '', name)
    if len(name) >= 28:
        name = name[0:28]
    return name

